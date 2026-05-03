"""Schedule parsing and loading.

Two entry points:

- ``parse_schedule(xlsx_path)`` — reads the master Excel and returns the flat
  session list. Used by the one-shot build script (``build_schedule.py``) and
  not by the running app.
- ``load_schedule(json_path)`` — reads the pre-parsed ``schedule.json`` that
  ships with the app. The Flask app uses this exclusively, so end users never
  need an Excel file present at runtime.

Excel quirk worth knowing about: double sessions like ``HRM-1 & 2: RK`` are
**merged cells** that span two consecutive time-slot columns (e.g. 14:30–16:00
+ 16:15–17:45 = 3 hours of class with a 15-min break in the middle). The
parser uses openpyxl directly so it can see merge ranges and emit one event
per time-slot the merge covers, with the same session label on both.
"""

from __future__ import annotations

import datetime
import json
import re
from typing import Any

SHEET_NAME = "Schedule"

# openpyxl is 1-indexed. Date is in column 2; the original spec referenced
# pandas 0-indexed columns (date = col 1), but we now read with openpyxl so
# everything below is in openpyxl's coordinate system.
DATE_COL = 2

# openpyxl column index → (start, end) for the five teaching slots.
# (Column 6, the 13:30–14:30 lunch slot, is intentionally absent.)
TIME_SLOTS: dict[int, tuple[str, str]] = {
    3: ("08:30", "10:00"),
    4: ("10:15", "11:45"),
    5: ("12:00", "13:30"),
    7: ("14:30", "16:00"),
    8: ("16:15", "17:45"),
}

# Order does not matter for re.match (it anchors at start), but listing the
# longer codes first is harmless and slightly more defensive.
KNOWN_CODES: list[str] = [
    "SAAPM", "EMABE", "CSLBA",
    "CDA", "BDM", "HRM", "CRF", "PRO", "IBC", "FRM", "CMF",
    "PM", "SM", "BE",
]


def _coerce_date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def _detect_course(cell: str) -> str | None:
    text = cell.strip()
    if not text:
        return None
    for code in KNOWN_CODES:
        if re.match(rf"^{code}[\-\s]", text):
            return code
    return None


def _build_merge_lookup(ws) -> dict[tuple[int, int], Any]:
    """Map every (row, col) inside a merged range to the top-left cell value.

    openpyxl stores merged-cell values only on the top-left cell; every other
    cell in the merge has ``value == None``. This lookup lets us treat each
    cell independently and still resolve the right value.
    """
    lookup: dict[tuple[int, int], Any] = {}
    for merged in ws.merged_cells.ranges:
        topleft = ws.cell(merged.min_row, merged.min_col).value
        if topleft is None:
            continue
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                lookup[(r, c)] = topleft
    return lookup


def parse_schedule(filepath: str) -> list[dict]:
    """Parse the master Excel into session dicts. Build-time use only."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb[SHEET_NAME]
    merge_lookup = _build_merge_lookup(ws)

    sessions: list[dict] = []
    for row_idx in range(1, ws.max_row + 1):
        date = _coerce_date(ws.cell(row_idx, DATE_COL).value)
        if date is None:
            continue

        for col, (start, end) in TIME_SLOTS.items():
            value = merge_lookup.get((row_idx, col))
            if value is None:
                value = ws.cell(row_idx, col).value
            if not isinstance(value, str):
                continue
            course = _detect_course(value)
            if course is None:
                continue
            sessions.append({
                "date": date,
                "day": date.strftime("%A"),
                "start": start,
                "end": end,
                "course": course,
                "session": value.strip(),
            })

    sessions.sort(key=lambda s: (s["date"], s["start"]))
    return sessions


def serialise_sessions(sessions: list[dict]) -> list[dict]:
    """Convert ``date`` objects to ISO strings so the list is JSON-safe."""
    return [
        {**s, "date": s["date"].isoformat()}
        for s in sessions
    ]


def load_schedule(json_path: str) -> list[dict]:
    """Load the pre-parsed schedule shipped with the app."""
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    sessions: list[dict] = []
    for s in raw:
        sessions.append({
            **s,
            "date": datetime.date.fromisoformat(s["date"]),
        })
    sessions.sort(key=lambda s: (s["date"], s["start"]))
    return sessions
